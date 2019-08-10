from . import assets
from .. import db
from ..models import Host, Capacity
from sqlalchemy import distinct, func, or_
from flask import jsonify, render_template, request, redirect, url_for, send_from_directory
from flask_login import login_required
from werkzeug.utils import secure_filename
from ..settings import ALLOWED_EXTENSIONS, UPLOAD_FOLDER, TEMPLATE_FOLDER, DOWNLOAD_FOLDER


# 主机详细信息
@assets.route("/info/<hostname>", methods=['GET', 'POST'])
@login_required
def get_host_info(hostname):
    host = Host.query.filter(Host.hostname == hostname).first()
    return render_template('assets_info.html', app="资产管理", action="资产详情", host=host)


# 更新主机信息
@assets.route("/update/<hostname>", methods=['GET', 'POST'])
@login_required
def update_host_info(hostname):
    if request.method == "GET":
        host_info = Host.query.filter(Host.hostname == hostname).first()
        return render_template('assets_update.html', app="资产管理", action="资产更新", host=host_info)
    elif request.method == "POST":
        m_host = request.form
        try:
            db.session.query(Host).filter(Host.hostname == hostname).update(m_host)
            db.session.commit()
            return render_template('assets_success.html', app="资产管理", action="资产更新", info='更新成功')
        except Exception as e:
            print(str(e))
            return render_template('assets_success.html', app="资产管理", action="资产更新", info='更新失败')


# 按业务平台分类获取节点信息
@assets.route("/get_nodes", methods=['GET', 'POST'])
@login_required
def get_nodes():
    cluste_type = db.session.query(distinct(Host.platform)).all()
    nodes = [x[0] for x in cluste_type]
    # ztree 一级节点
    z_nodes = []
    for i, node in enumerate(nodes):
        cluste_temps = db.session.query(distinct(Host.cluster)).filter(Host.platform == node).order_by(
            Host.cluster).all()
        clusters = [x[0] for x in cluste_temps]
        # 二级节点
        childrens = [{'name': cluster} for cluster in clusters]
        p1_data = {
            # 'name': node + " ({})".format(len(clusters)),
            'name': node,
            'open': 'true' if i == 0 else 'false',
            'children': childrens
        }
        z_nodes.append(p1_data)
    return jsonify(z_nodes)


# 获取对应集群下的主机清单，集群名称通过post参数传递
@assets.route("/get_hosts", methods=['GET', 'POST'])
def get_hosts():
    cluester = request.form.get('cluster')
    hosts_temp = Host.query.filter(Host.cluster == cluester).all()
    # hosts = [x.hostname for x in hosts_temp]
    result = []
    for i, host in enumerate(hosts_temp):
        # result.append({
        #     'id': i,
        #     'name': host.hostname,
        #     'ip': host.local_ip,
        #     'device_type': host.device_type,
        #     'device_model': host.device_model,
        #     'engine_room': host.engine_room,
        #     'frame_number': host.frame_number
        # })
        result.append(host.to_json())
    # result = [{'id': i, 'name': x.hostname} for i, x in enumerate(hosts_temp)]
    return jsonify(result)


# 资产列表主页
@assets.route("/asset", methods=['GET', 'POST'])
@login_required
def asset():
    return render_template('assets_list.html', app="资产管理", action="资产列表")


# 资产统计
@assets.route("/dashboard", methods=['GET', 'POST'])
@login_required
def dashboard():
    # 按业务平台分类汇总数量
    result = db.session.query(Host.platform, func.count('*').label("count")).filter(Host.status=="在网").group_by(Host.platform).all()
    platforms = [x[0] for x in result]
    platforms_counts = [x[1] for x in result]
    device_nums = sum(platforms_counts)
    # 按机房分类获取数量
    result_jf = db.session.query(Host.engine_room, func.count('*').label("count")).filter(Host.status=="在网").group_by(Host.engine_room).all()
    engine_rooms = [x[0] for x in result_jf]
    engine_rooms_values= [{"value": v, "name": k} for k, v in result_jf]
    # 按设备类型分类获取数量
    result_type = db.session.query(Host.device_type, func.count('*').label("count")).filter(Host.status=="在网").group_by(Host.device_type).all()
    device_types = [x[0] for x in result_type]
    device_types_nums = [x[1] for x in result_type]
    # device_types_values = [{"value": v, "name": k} for k, v in result_type]
    # 统计操作系统
    linux_nums = db.session.query(func.count('*').label("count")).filter(
        or_(Host.os_version.like("linux%"), Host.os_version.like("redhat%"))).all()
    hpux_nums = db.session.query(func.count('*').label("count")).filter(
        or_(Host.os_version.like("HP%"), Host.os_version.like("hp%"))).all()
    aix_nums = db.session.query(func.count('*').label("count")).filter(
        or_(Host.os_version.like("AIX%"), Host.os_version.like("aix%"))).all()
    return render_template('assets_dashboard.html', app="资产管理", action="资产统计",  **locals())


# 创建资产
@assets.route('/create', methods=["GET", "POST"])
@login_required
def create():
    if request.method == "GET":
        # 获取platform cluster参数值，作为创建资产的默认值
        platform = request.args.get('platform', '')
        cluster = request.args.get('cluster', '')
        return render_template('assets_create.html', app="资产管理", action="创建资产",
                               platform=platform, cluster=cluster)
    elif request.method == "POST":
        # 获取提交的表单数据，转为dict
        _host = request.form
        host = _host.to_dict()
        try:
            add_host = Host(**host)
            db.session.add(add_host)
            db.session.commit()
            info = "创建成功"
        except Exception as e:
            info = str(e)
        return render_template('assets_success.html', app="资产管理", action="创建资产", info=info)


# 删除设备
@assets.route('/delete', methods=["GET", "POST"])
@login_required
def delete():
    hostname = request.form.get('id')
    try:
        to_delete = Host.query.filter(Host.hostname == hostname).first()
        db.session.delete(to_delete)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    return jsonify(result)


# 批量删除设备
@assets.route('/multi_delete', methods=["GET", "POST"])
@login_required
def multi_delete():
    _hosts = request.form.get('hosts')
    _hosts_temp = _hosts.split(',')
    # 去除checkbox选择的无效选项
    _hosts_list = [x for x in _hosts_temp if x != '' and x != 'on']
    if _hosts_list:
        for hostname in _hosts_list:
            to_delete = Host.query.filter(Host.hostname == hostname).first()
            db.session.delete(to_delete)
        db.session.commit()
        return "success"
    else:
        return "fail"


# 全局缓存字典，临时存放变量
assets_cache = {}
# 资产批量更新
@assets.route('/multi_update', methods=["GET", "POST"])
@login_required
def multi_update():
    if request.method == "GET":
        _hosts = request.args.get('hosts')
        _hosts_temp = _hosts.split(',')
        # 去除checkbox选择的无效选项
        _hosts_list = [x for x in _hosts_temp if x != '' and x != 'on']
        if _hosts_list:
            assets_cache['hosts'] = _hosts_list
            return render_template('assets_multi_update.html', app="资产管理", action="资产更新", hosts=_hosts_list)
        else:
            return redirect(url_for('assets.asset'))
    elif request.method == "POST":
        hosts = assets_cache.get('hosts')
        # 清空
        assets_cache['hosts'] = None
        _update_datas = request.form.to_dict()
        update_datas = {k: v for k, v in _update_datas.items() if v}
        for host in hosts:
            db.session.query(Host).filter(Host.hostname == host).update(update_datas)
        db.session.commit()
        return render_template('assets_success.html', app="资产管理", action="资产更新", info='更新成功')


# 资产导入模板允许文件
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# 下载导入资产的模板
@assets.route('/down_templates/<filename>')
@login_required
def down_templates(filename):
    # filename = 'asset_data.xlsx'
    return send_from_directory(TEMPLATE_FOLDER, filename=filename, as_attachment=True)


# 导入资产
@assets.route('/fm_import', methods=["GET", "POST"])
@login_required
def fm_import():
    import os
    if request.method == 'POST':
        # 请求中无文件
        if 'file' not in request.files:
            return jsonify({'flag': 'fail', 'msg': '上传文件失败'})
        file = request.files['file']
        # 文件名为空
        if file.filename == '':
            return jsonify({'flag': 'fail', 'msg': '文件名错误'})
        # 文件符合要求
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            # 调用入库函数
            nums = load_to_db(UPLOAD_FOLDER + filename, data_type="asset")
            return jsonify({'flag': 'success', 'msg': '导入成功' + str(nums)})
        else:
            return jsonify({'flag': 'fail', 'msg': '上传文件格式错误'})


# 导入资产入库
def load_to_db(filename, data_type):
    from openpyxl import load_workbook
    nums = 0
    # 与数据中字段一致
    capacity_names = ["platform", "cluster", "s_capacity", "h_capacity", "status", "info"]
    asset_names = ["platform", "cluster", "hostname", "device_type", "manufacturer", "device_model", "serial",
                  "account", "version", "software_version", "local_ip", "nat_ip", "os_version", "engine_room",
                  "frame_number", "power_frame_number", "net_time", "s_period", "h_period", "power", "status"]
    infos = {
        "capacity": {
            "cols": capacity_names,
            "sheet": "capacity",
            "model": Capacity
        },
        "asset": {
            "cols": asset_names,
            "sheet": "asset_data",
            "model": Host
        }
    }
    try:
        wb = load_workbook(filename)
        # 读取导入excel文件中的sheet表
        ws = wb[infos[data_type].get("sheet")]
        # 获取列数
        cols = len(infos[data_type].get("cols"))
        db_model = infos[data_type].get("model")
        # 获取行数
        rows = ws.max_row
        dbs = []
        # 读取excel中每行数据，构造成dict, 方便入库
        for i in range(2, rows + 1):
            data = []
            for j in range(1, cols + 1):
                value = ws.cell(i, j).value
                if not value:
                    value = ""
                data.append(value)
            item = {k: v for k, v in zip(infos[data_type].get("cols"), data)}
            dbs.append(db_model(**item))
        # 批量添加
        db.session.add_all(dbs)
        db.session.commit()
        nums = len(dbs)
    except Exception as e:
        print(str(e))
    finally:
        # 最后删除上传的文件
        import os
        os.remove(filename)
    return nums


# 导出结果
@assets.route('/unload', methods=["GET", "POST"])
@login_required
def unload_excel():
    import json, datetime, os, pandas
    # 清空目录
    try:
        filelist = os.listdir(DOWNLOAD_FOLDER)
        for file in filelist:
            os.remove(DOWNLOAD_FOLDER+file)
    except Exception as e:
        print(str(e))
    temp = request.form.get('data')
    if temp:
        # 需转为json, 不然读取时为字符串
        data_list = json.loads(temp)
        datas = []
        for x in data_list:
            if x[-1] == "动作"or x[-1] == "更新\n删除" or x[-1] == "更新删除":
                datas.append(x[1:-1])
            else:
                datas.append(x)
        print(datas)
        filename = datetime.datetime.now().strftime("%Y%m%d%H%M%S") + ".xlsx"
        # 转成pandas  DataFrame格式
        df = pandas.DataFrame(datas[1:], columns=datas[0])
        # 导出excel
        df.to_excel(DOWNLOAD_FOLDER + filename, index=True)
        # 返回生成的excel文件名
        result = {
            "flag": "success",
            "file": filename
        }
    else:
        result = {"flag": "fail",
                  "file": ""
                  }
    return jsonify(result)


# 下载导出的文件
@assets.route('/download/<filename>')
@login_required
def download_file(filename):
    return send_from_directory(DOWNLOAD_FOLDER, filename=filename, as_attachment=True)


# 资产查询
@assets.route('/query', methods=["GET", "POST"])
@login_required
def query():
    platforms = db.session.query(distinct(Host.platform)).all()
    device_types = db.session.query(distinct(Host.device_type)).all()
    engine_rooms = db.session.query(distinct(Host.engine_room)).all()
    clusters = db.session.query(distinct(Host.cluster)).all()
    platform = [x[0] for x in platforms]
    device_type = [x[0] for x in device_types]
    engine_room = [x[0] for x in engine_rooms]
    cluster = [x[0] for x in clusters]
    return render_template('assets_query.html', app="资产管理", action="资产查询",
                           platform=platform, device_type=device_type, engine_room=engine_room, cluster=cluster)

# 返回查询结果
@assets.route('/get_query_hosts', methods=["GET", "POST"])
@login_required
def get_query_hosts():
    temp = ["选择平台", "选择机房", "选择网元", "选择设备类型"]
    pt = request.form.get('pt')
    jf = request.form.get('jf')
    jq = request.form.get('jq')
    lx = request.form.get('lx', '')
    # 构造查询条件
    kw_temp = {
        "platform": pt if pt and pt not in temp else None,
        "engine_room": jf if jf and jf not in temp else None,
        "cluster": jq if jq and jq not in temp else None,
        "device_type": lx if lx and lx not in temp else None
    }
    # 去除查询条件为 None 的选项
    kw = {k: v for k, v in kw_temp.items() if v}
    # 查询结果
    hosts = db.session.query(Host).filter_by(**kw).all()
    datas = []
    # host数据转成json类型，前端解析
    for host in hosts:
        datas.append(host.to_json())
    result = {
        "flag": "success",
        "hosts": datas,
        "counts": len(datas)
    }
    return jsonify(result)


# 搜索
@assets.route('/search', methods=["GET", "POST"])
@login_required
def search():
    name = request.form.get('name')
    # 模糊搜索
    name = '%' + name.upper() + '%'
    datas = []
    hosts = db.session.query(Host).filter(or_(Host.hostname.like(name), Host.cluster.like(name))).all()
    if hosts:
        # host数据转成json类型，前端解析
        for host in hosts:
            datas.append(host.to_json())
        result = {
            "flag": "success",
            "hosts": datas,
            "counts": len(datas)
        }
    else:
        result = {
            "flag": "fail"
        }
    return jsonify(result)


# 系统容量
@assets.route('/capacity', methods=["GET", "POST"])
@login_required
def capacity():
    if request.method == "GET":
        platforms = db.session.query(distinct(Capacity.platform)).all()
        platform = [x[0] for x in platforms]
        datas = Capacity.query.all()
        return render_template('assets_capacity.html', app="资产管理", action="系统容量", datas=datas, platform=platform)
    elif request.method == "POST":
        platform = request.form.get('platform')
        hosts_temp = Capacity.query.filter(Capacity.platform == platform).all()
        result = []
        for host in hosts_temp:
            result.append(host.to_json())
        return jsonify({"flag": "success", "datas": result})


# 根据id获取容量信息
@assets.route('/get_capacity_info', methods=["GET", "POST"])
@login_required
def get_capacity_info():
    item_id = request.form.get('id')
    item_info = db.session.query(Capacity).filter(Capacity.id == item_id).one()
    result = {
        "flag": "success",
        "item_info": item_info.to_json()
    }
    return jsonify(result)


# 添加容量信息
@assets.route('/create_capacity_info', methods=["GET", "POST"])
@login_required
def create_capacity_info():
    if request.method == "GET":
        return render_template('assets_capacity_create.html', app='资产管理', action='系统容量')
    else:
        item = request.form.to_dict()
        add_item = Capacity(**item)
        db.session.add(add_item)
        db.session.commit()
        return redirect(url_for('assets.capacity'))


# 修改容量信息
@assets.route('/modify_capacity_info', methods=["GET", "POST"])
@login_required
def modify_capacity_info():
    # 获取前端提交的json数据
    datas = request.get_json()
    try:
        item_id = datas.get('id')
        db.session.query(Capacity).filter(Capacity.id == item_id).update(datas)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    # request中要求的数据格式为json，为其他会导致前端执行success不成功
    return jsonify(result)


# 删除容量信息
@assets.route('/delete_capacity_info', methods=["GET", "POST"])
@login_required
def delete_capacity_info():
    item_id = request.form.get('id')
    try:
        to_delete = Capacity.query.filter(Capacity.id == item_id).first()
        db.session.delete(to_delete)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    return jsonify(result)


# 导入容量
@assets.route('/ca_import', methods=["GET", "POST"])
@login_required
def ca_import():
    import os
    if request.method == 'POST':
        # 请求中无文件
        if 'file' not in request.files:
            return jsonify({'flag': 'fail', 'msg': '上传文件失败'})
        file = request.files['file']
        # 文件名为空
        if file.filename == '':
            return jsonify({'flag': 'fail', 'msg': '文件名错误'})
        # 文件符合要求
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            # 调用入库函数
            nums = load_to_db(UPLOAD_FOLDER + filename, data_type="capacity")
            return jsonify({'flag': 'success', 'msg': '导入成功' + str(nums)})
        else:
            return jsonify({'flag': 'fail', 'msg': '上传文件格式错误'})
