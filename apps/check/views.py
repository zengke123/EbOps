from . import check
from .. import db, scheduler
from ..models import CheckHistory, CheckHost, ChenkJobs
from sqlalchemy import distinct
from flask import render_template, request, jsonify, send_from_directory
from flask_login import login_required, current_user
from .exts import req_zjlj, req_pllj, test_check, add_job_scheduler
from werkzeug.utils import secure_filename
from ..settings import ALLOWED_EXTENSIONS, UPLOAD_FOLDER, TEMPLATE_FOLDER, CHECK_DOWNLOAD_FOLDER
# 例检状态全局变量
status = {}


@check.route("/info", methods=['GET', 'POST'])
@login_required
def info():
    return render_template('check_list.html', app="自动例检", action="主机例检")


@check.route("/history", methods=['GET', 'POST'])
@login_required
def history():
    # 获取get请求传过来的页数,没有传参数，默认为1
    page = int(request.args.get('page', 1))
    date = request.args.get('date')
    if date and date != 'None':
        paginate = CheckHistory.query.filter(CheckHistory.checktime.like(date+'%')).order_by(CheckHistory.id.asc()).paginate(page, per_page=10, error_out=False)
    else:
        paginate = CheckHistory.query.order_by(CheckHistory.id.asc()).paginate(page, per_page=10, error_out=False)
    datas = paginate.items
    return render_template('check_history.html', app="自动例检", action="例检记录", paginate=paginate, datas=datas, date=date)


@check.route("/lj", methods=['GET', 'POST'])
@login_required
def check_host():
    _args = request.form.get('item_id', None)
    names = _args.split('_')
    print(names)
    if len(names) == 2:
        host_type = names[0]
        hostname = names[1]
        api_name = {'type':host_type, 'name':hostname}
        zjlj_task = req_zjlj.delay(api_name, hostname, current_user.username)
        # zjlj_task = test_check.delay(lj_type=api_name, name=hostname, operator=current_user.username)
        return jsonify({'flag': 'success', 'desc': '任务已添加[id:{}]'.format(zjlj_task.id)})
    else:
        return jsonify({'flag': 'fail', 'desc': '参数错误'})
    # return render_template('check_confirm.html', app="自动例检", action="例检确认",
    #                        hostname=hostname, host_type=host_type)


@check.route("/multi_check", methods=['GET', 'POST'])
@login_required
def multi_check():
    _args = request.form.get('hosts', None)
    print(_args)
    if _args:
        _hosts = _args.split(',')
        hosts = ["'" + str(x) + "'" for x in _hosts if x != ""]
        print(hosts)
        names = ",".join(hosts)
        print(names)
        api_name = {"type": "pl", "name": names}
        print(api_name)
        zjlj_task = req_pllj.delay(api_name, _args, current_user.username)
        return jsonify({'flag': 'success', 'desc': '任务已添加[id:{}]'.format(zjlj_task.id)})
    else:
        return jsonify({'flag': 'fail', 'desc': '参数错误'})


# 下载例检报告
@check.route('/download/<file>')
def download(file):
    filepath = CHECK_DOWNLOAD_FOLDER + str(file) + '/'
    filename = str(file) + ".tar.gz"
    return send_from_directory(filepath, filename, as_attachment=True)


# 例检配置
@check.route('/config', methods=['GET', 'POST'])
@login_required
def config():
    if request.method == "GET":
        return render_template('check_config.html', app="自动例检", action="例检配置")
    elif request.method == "POST":
        data = request.form.to_dict()
        print(data)
        add_host = CheckHost(**data)
        db.session.add(add_host)
        db.session.commit()
        info = "创建成功"
        return render_template('check_set_success.html', app="自动例检", action="例检配置", info=info)


# 批量导入例检配置
@check.route('/config/load', methods=['GET', 'POST'])
@login_required
def load_config():
    return render_template('check_load.html', app="自动例检", action="例检配置")


# 下载导入例检主机模板
@check.route('/down_templates')
@login_required
def down_templates():
    filename = 'check_data.xlsx'
    return send_from_directory(TEMPLATE_FOLDER, filename=filename, as_attachment=True)


# 资产导入模板允许文件
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# 导入例检主机
@check.route('/fm_import', methods=["GET", "POST"])
@login_required
def fm_import():
    import os
    if request.method == 'POST':
        # 请求中无文件
        if 'file' not in request.files:
            return jsonify({'flag': 'fail', 'msg': '上传文件失败'})
        file = request.files['file']
        # 文件名为空
        if file.filename == '':
            return jsonify({'flag': 'fail', 'msg': '文件名错误'})
        # 文件符合要求
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            # 调用入库函数
            nums = load_to_db(UPLOAD_FOLDER + filename)
            return jsonify({'flag': 'success', 'msg': '导入成功' + str(nums)})
        else:
            return jsonify({'flag': 'fail', 'msg': '上传文件格式错误'})


# 导入资产入库
def load_to_db(filename):
    from openpyxl import load_workbook
    nums = 0
    # 与数据中字段一致
    host_names = ["os", "pt", "jq", "zj", "zh", "jctype", "jcnum", "ip"]
    infos = {
        "cols": host_names,
        "sheet": "hosts"
        }
    try:
        wb = load_workbook(filename)
        # 读取导入excel文件中的sheet表
        ws = wb[infos.get("sheet")]
        # 获取列数
        cols = len(infos.get("cols"))
        # 获取行数
        rows = ws.max_row
        dbs = []
        # 读取excel中每行数据，构造成dict, 方便入库
        for i in range(2, rows + 1):
            data = []
            for j in range(1, cols + 1):
                value = ws.cell(i, j).value
                if not value:
                    value = ""
                data.append(value)
            item = {k: v for k, v in zip(infos.get("cols"), data)}
            dbs.append(CheckHost(**item))
        # 批量添加
        db.session.add_all(dbs)
        db.session.commit()
        nums = len(dbs)
    except Exception as e:
        print(str(e))
    finally:
        # 最后删除上传的文件
        import os
        os.remove(filename)
    return nums


# 删除设备
@check.route('/delete', methods=["GET", "POST"])
@login_required
def delete():
    hostname = request.form.get('id')
    try:
        to_deletes = CheckHost.query.filter(CheckHost.zj == hostname).all()
        for to_delete in to_deletes:
            db.session.delete(to_delete)
        db.session.commit()
        result = {"flag": "success"}
    except Exception as e:
        print(str(e))
        result = {"flag": "fail"}
    return jsonify(result)


# 批量删除设备
@check.route('/multi_delete', methods=["GET", "POST"])
@login_required
def multi_delete():
    _hosts = request.form.get('hosts')
    _hosts_temp = _hosts.split(',')
    # 去除checkbox选择的无效选项
    _hosts_list = [x for x in _hosts_temp if x != '' and x != 'on']
    if _hosts_list:
        for hostname in _hosts_list:
            to_deletes = CheckHost.query.filter(CheckHost.zj == hostname).all()
            for to_delete in to_deletes:
                db.session.delete(to_delete)
        db.session.commit()
        return "success"
    else:
        return "fail"


# 例检任务
@check.route('/jobs')
@login_required
def jobs():
    all_jobs = db.session.query(ChenkJobs).all()
    run_jobs = ChenkJobs.query.filter(ChenkJobs.status == 1).all()
    return render_template("check_jobs.html", app="自动例检", action="例检任务", run_jobs=run_jobs, all_jobs=all_jobs)


@check.route('/jobs/add', methods=['GET', 'POST'])
@login_required
def add_job():
    name = request.form.get('name')
    content = request.form.get('content')
    cron_time = request.form.get('cron_time')
    host = content.split(",")
    _hosts_data = db.session.query(CheckHost.zj).all()
    _hosts = [x[0] for x in _hosts_data]
    _clusters_data = db.session.query(distinct(CheckHost.jq)).all()
    _clusters = [x[0] for x in _clusters_data]
    args = []
    for x in host:
        if x in _hosts:
            x = 'zj_' + x
        elif x in _clusters:
            x = 'jq_' + x
        else:
            return jsonify({'result': 'fail', 'error': x + "不存在"})
        args.append(x)
    need_add_job = ChenkJobs(name=name, content='|'.join(args), cron_time=cron_time)
    if cron_time.count(",") != 4:
        return jsonify({'result': 'fail', 'error': "执行时间Cron格式错误"})
    db.session.add(need_add_job)
    try:
        db.session.commit()
        return jsonify({'result': 'success', 'error': None})
    except:
        return jsonify({'result': 'fail', 'error': '数据库错误'})


@check.route('/jobs/pause', methods=['GET', 'POST'])
@login_required
def pause_job():
    job_id = request.form.get('job_id')
    current_job = ChenkJobs.query.filter(ChenkJobs.id == int(job_id)).first()
    if current_job.status == 1:
        scheduler.pause_job(id=job_id)
        current_job.status = 0
        db.session.commit()
        print("任务【{}】【{}】已从scheduler队列暂停".format(current_job.id, current_job.name))
        return "success"
    else:
        return "fail"


@check.route('/jobs/active', methods=['GET', 'POST'])
@login_required
def active_job():
    # 未初始化的任务需要先进行添加
    job_id = request.form.get('job_id')
    current_job = ChenkJobs.query.filter(ChenkJobs.id == int(job_id)).first()
    if current_job:
        try:
            scheduler.resume_job(id=int(job_id))
            current_job.status = 1
            db.session.commit()
        except:
            _args = (current_job.content,)
            add_job_scheduler(scheduler, job_id=current_job.id, job_cron=current_job.cron_time, args=_args)
            current_job.status = 1
            db.session.commit()
        print("任务【{}】【{}】已从scheduler队列激活".format(current_job.id, current_job.name))
        return "success"
    else:
        return "fail"


@check.route('/jobs/remove', methods=['GET', 'POST'])
@login_required
def remove_job():
    # 任务删除,正在运行或已暂停的任务，需要从任务队列中清除
    job_id = request.form.get('job_id')
    current_job = ChenkJobs.query.filter(ChenkJobs.id == int(job_id)).first()
    try:
        scheduler.delete_job(id=job_id)
        print("任务【{}】【{}】已从scheduler队列删除".format(current_job.id, current_job.name))
    except:
        print("任务【{}】【{}】未在scheduler队列初始化,已删除".format(current_job.id, current_job.name))
    try:
        db.session.delete(current_job)
        db.session.commit()
        return "success"
    except:
        return "fail"